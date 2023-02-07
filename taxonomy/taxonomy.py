from collections import defaultdict
import openpyxl as xls
import json

def rdict():
    return defaultdict(rdict)

class Taxonomy:

    def __init__(self):
        self._props = {'JP': {'edinet': self._edinet},
                        'US': {'gaap': self._us_sec}}
        self._edinet_props = {'jp_gaap': {'tab': 2, 'prefix': 'H', 'tag': 'I', 'description': 'B'},
                              'ifrs': {'tab': 3, 'prefix': 'G', 'tag': 'H', 'description': 'B'}}
        
    def load(self, mode='ifrs'):
        return self._edinet(mode)

    def _edinet(self, file_path):

        def check_standard(wb):
            mode = ''
            if wb[wb.sheetnames[0]]['A'][0].value == '国際会計基準タクソノミ項目リスト 目次':
                mode = 'ifrs'
            if wb[wb.sheetnames[0]]['A'][0].value == '勘定科目リスト 目次':
                mode = 'jp_gaap'
            return self._edinet_props[mode] 

        wb = xls.load_workbook(file_path)
        #print(wb.sheetnames[2:])

        edp = check_standard(wb)
 
        d = rdict()
        k = ''
        tab = wb[wb.sheetnames[edp['tab']]]
        for col in zip(tab[edp['description']], tab[edp['prefix']], tab[edp['tag']]):
            if col[0].fill.bgColor.rgb != '00000000':
                k = col[0].value
                continue
            prefix = col[1].value
            tag = col[2].value
            d[prefix][tag] = {'category': k, 'description': col[0].value}
        
        print(json.dumps(d))

    def _us_sec(self, mode):
        pass
